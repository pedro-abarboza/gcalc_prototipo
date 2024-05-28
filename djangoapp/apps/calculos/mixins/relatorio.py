import openpyxl
from django.http import HttpResponse
from apps.calculos.models import Calculos, VerbasCalc
from apps.processos.models import Processos, Reclamadas, Reclamantes


class RelatorioMixin:

    def get_relatorio(self, context):    
        colunas_id = ''
        if 'processo' in self.request.GET:
            context['processo_id'] = self.request.GET['processo']
        if 'reclamante' in self.request.GET and self.request.GET['reclamante']:
            context['reclamante_id'] = self.request.GET['reclamante']
        if 'reclamada' in self.request.GET and self.request.GET['reclamada']:
            context['reclamada_id'] = self.request.GET['reclamada']
        if 'colunas[]' in self.request.GET and self.request.GET['colunas[]']:
            context['colunas_id'] = colunas_id = self.request.GET.getlist('colunas[]')
        
        kwargs_colunas = {'calculo_id__in':list(self.queryset.values_list('id',flat=True))}
        if colunas_id:
            kwargs_colunas['id__in'] = colunas_id

        context['title'] = self.get_title()
        context['card_title'] = 'Listagem'
        context['subtitle'] = 'Aqui você tem a lista de todos os calculos cadastrados.'
        context['processos'] = list(Processos.objects.all().distinct().values_list('id', 'n_processo'))
        context['reclamadas'] = list(Reclamadas.objects.all().distinct().values_list('id', 'nome'))
        context['reclamantes'] = list(Reclamantes.objects.all().distinct().values_list('id', 'nome'))
        context['colunas'] = list(VerbasCalc.objects.all().values_list('id', 'descricao')
            .exclude(descricao='Total').distinct('descricao').order_by('descricao'))
        context['colunas_tabela'] = list(VerbasCalc.objects
            .filter(**kwargs_colunas)
            .exclude(descricao='Total').distinct('descricao').order_by('descricao'))
        return context
    
    def get_export(self, context):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Relatório de Calculos G-Calc'

        colunas_tabela = context['colunas_tabela']
        calculos = self.queryset

        colunas_ini = [
            ('Arquivo','n_arquivo'),
            ('D.Liq.', 'get_dt_liquidacao'),
            ('ID','id'),
            ('Nº Processo','get_processo'),
            ('Reclamada','reclamada'),
            ('Reclamante','reclamante')
        ]
        col = 1
        for column_title in colunas_ini:
            cell = sheet.cell(row=1, column=col)
            cell.value = column_title[0]
            col+=1

        # Adicionar cabeçalhos à primeira linha
        for col_num, column_title in enumerate(colunas_tabela, start=1):
            cell = sheet.cell(row=1, column=col_num+col)
            cell.value = column_title.descricao

        # Adicionar alguns dados
        data = []
        for calculo in calculos:
            linha = []
            for coluna in colunas_ini:
                linha.append(getattr(calculo, coluna[1]))

            for coluna in colunas_tabela:
                linha.append(calculo.get_coluna(coluna.descricao))

            data.append(linha)
        
        for row_num, row_data in enumerate(data, start=2):
            for col_num, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Salvar o workbook no response
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Relatório de Cálculos G-Calc.xlsx"'
        wb.save(response)
        return response